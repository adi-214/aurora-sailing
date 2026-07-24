#!/usr/bin/env python3
"""
find_emails.py — bulk work-email finder driven by one Excel workbook.

    Companies sheet   -> what to look up + where results are written
    API Keys sheet    -> one row per key
    Settings sheet    -> a few knobs

    pip install openpyxl requests
    python find_emails.py --workbook Aurora_Email_Finder.xlsx

    --check-keys   ask each provider what your key can actually do (free, 0 credits)
    --dry-run      no network, fills plausible fake data to test the plumbing
    --limit 5      only the first 5 pending rows
    --redo         re-run rows that already have an email
    --verbose      print the raw error body whenever a call fails

Design rule: a bad row never kills the run, and a bad request never kills a key.
Only a 401/402/403, or a 429 that says "usage limit reached", retires a key.
"""

import argparse
import datetime as dt
import re
import sys
import time

try:
    import requests
except ImportError:
    requests = None

from openpyxl import load_workbook

HUNTER = "https://api.hunter.io/v2"
APOLLO = "https://api.apollo.io/api/v1"

# Hunter's free plan rejects limit + offset > 10 with "400 pagination_error".
# 10 is also the most emails per domain a free plan will ever return, and a
# domain-search costs the same 1 credit for anything from 1 to 10 results.
HUNTER_PAGE = 10

SHEET_COMPANIES, SHEET_KEYS, SHEET_SETTINGS = "Companies", "API Keys", "Settings"

COL = {"id": 1, "org": 2, "domain": 3, "first": 4, "last": 5, "role": 6, "mode": 7,
       "out_domain": 8, "out_email": 9, "out_name": 10, "out_position": 11,
       "out_confidence": 12, "out_provider": 13, "out_status": 14,
       "out_other": 15, "out_run": 16}
FIRST_DATA_ROW = 3


# --------------------------------------------------------------------------- #
#  Errors: the distinction the old script got wrong
# --------------------------------------------------------------------------- #
class ApiError(Exception):
    """One failed call. `kind` decides what we do about it."""

    def __init__(self, provider, status, code="", detail="", kind="request"):
        self.provider, self.status = provider, status
        self.code, self.detail, self.kind = code, detail, kind
        super().__init__(f"HTTP {status} {code} {detail}".strip())


def classify(status, code):
    """key = this key is unusable. rate = slow down. request = we asked wrong."""
    if status in (401, 402, 403):
        return "key"
    if status == 429:
        return "key" if code in ("usage_limit_reached", "plan_limit_reached") else "rate"
    if status >= 500:
        return "server"
    return "request"          # 400, 404, 422 ... our problem, not the key's


def warn(where, e):
    """Report a failed call, unless it is just 'that key already retired'."""
    if e.code != "no_key":
        print(f"        {where}: {e}")


# --------------------------------------------------------------------------- #
#  Keys
# --------------------------------------------------------------------------- #
class Key:
    def __init__(self, provider, value, priority, note=""):
        self.provider = str(provider).strip().lower()
        self.value = str(value).strip()
        self.priority = priority if priority is not None else 999
        self.note, self.dead, self.calls = note or "", False, 0

    @property
    def tail(self):
        return self.value[-4:] if len(self.value) >= 4 else self.value


class Keys:
    def __init__(self, keys):
        self.all = sorted(keys, key=lambda k: k.priority)

    def live(self, provider):
        return next((k for k in self.all if k.provider == provider and not k.dead), None)

    def retire(self, key, why):
        key.dead = True
        print(f"        key ...{key.tail} ({key.provider}) retired — {why}")

    def by_provider(self):
        out = {}
        for k in self.all:
            out.setdefault(k.provider, []).append(k)
        return out


# --------------------------------------------------------------------------- #
#  Providers
# --------------------------------------------------------------------------- #
def _body(resp):
    try:
        return resp.json()
    except Exception:
        return {}


def hunter(keys, path, params, verbose=False, method="get"):
    """Call a Hunter endpoint. Rotates keys only on a genuine key fault."""
    while True:
        key = keys.live("hunter")
        if not key:
            raise ApiError("hunter", 0, "no_key", "no usable Hunter key left", "key")
        if method == "post":
            r = requests.post(f"{HUNTER}{path}", json=params,
                              params={"api_key": key.value}, timeout=30)
        else:
            r = requests.get(f"{HUNTER}{path}",
                             params={**params, "api_key": key.value}, timeout=30)
        key.calls += 1
        if r.status_code == 200:
            return _body(r).get("data", {}) or {}

        body = _body(r)
        err = (body.get("errors") or [{}])[0]
        code, detail = err.get("id", ""), err.get("details", "")
        if verbose:
            print(f"        raw: {r.status_code} {body}")
        kind = classify(r.status_code, code)

        if kind == "rate":
            time.sleep(2.0)
            continue
        if kind == "key":
            keys.retire(key, f"HTTP {r.status_code} {code} {detail}".strip())
            continue
        raise ApiError("hunter", r.status_code, code, detail, kind)


def apollo(keys, path, payload, verbose=False, params=None):
    """POST an Apollo endpoint. Same rules. Some endpoints want URL query params."""
    while True:
        key = keys.live("apollo")
        if not key:
            raise ApiError("apollo", 0, "no_key", "no usable Apollo key left", "key")
        headers = {"X-Api-Key": key.value, "Content-Type": "application/json",
                   "Cache-Control": "no-cache"}
        r = requests.post(f"{APOLLO}{path}", json=payload, headers=headers,
                          params=params or {}, timeout=30)
        key.calls += 1
        if r.status_code == 200:
            return _body(r)

        body = _body(r)
        detail = str(body.get("error") or body.get("message") or "")[:120]
        if verbose:
            print(f"        raw: {r.status_code} {body}")
        kind = classify(r.status_code, "")

        if kind == "rate":
            time.sleep(2.0)
            continue
        if kind == "key":
            keys.retire(key, f"HTTP {r.status_code} {detail} "
                             f"(Apollo gates API access behind paid plans)")
            continue
        raise ApiError("apollo", r.status_code, "", detail, kind)


def check_keys(keys, verbose=False):
    """Ask each provider what the key can do. Hunter's /account costs 0 credits."""
    for k in keys.all:
        label = f"{k.provider:<7} ...{k.tail}"
        try:
            if k.provider == "hunter":
                r = requests.get(f"{HUNTER}/account", params={"api_key": k.value}, timeout=30)
                if r.status_code != 200:
                    err = (_body(r).get("errors") or [{}])[0]
                    print(f"  {label}  DEAD — HTTP {r.status_code} "
                          f"{err.get('id', '')} {err.get('details', '')}")
                    continue
                d = _body(r).get("data", {}) or {}
                s = (d.get("requests", {}) or {}).get("searches", {}) or {}
                print(f"  {label}  OK — plan '{d.get('plan_name', '?')}', "
                      f"searches used {s.get('used', '?')} of {s.get('available', '?')}")
            else:
                r = requests.get(f"{APOLLO}/auth/health",
                                 headers={"X-Api-Key": k.value}, timeout=30)
                d = _body(r)
                if r.status_code == 200 and d.get("is_logged_in"):
                    print(f"  {label}  OK — key authenticates")
                else:
                    print(f"  {label}  DEAD — HTTP {r.status_code} {str(d)[:90]}")
            if verbose:
                print(f"        raw: {r.status_code} {_body(r)}")
        except Exception as e:
            print(f"  {label}  could not reach provider: {e}")


# --------------------------------------------------------------------------- #
#  Picking the best email out of a domain search
# --------------------------------------------------------------------------- #
def best_email(emails, role_hint):
    if not emails:
        return None, []
    hint = (role_hint or "").lower().strip()

    def score(e):
        s = e.get("confidence") or 0
        if hint:
            hay = " ".join(str(e.get(f) or "") for f in
                           ("position", "department", "seniority")).lower()
            if hint in hay:
                s += 1000
        if e.get("first_name") or e.get("last_name"):
            s += 5                      # a named person beats a generic inbox
        return s

    ranked = sorted(emails, key=score, reverse=True)
    return ranked[0], [e["value"] for e in ranked[1:] if e.get("value")]


def blank(domain="", status="not_found", note=""):
    return {"status": status, "provider": "", "confidence": "", "email": "",
            "name": "", "position": "", "domain": domain, "other": note}


STOPWORDS = {"the", "and", "of", "inc", "ltd", "pty", "limited", "co", "company",
             "incorporated", "group", "australia", "australian", "&"}


def name_tokens(s):
    return set(re.sub(r"[^a-z0-9 ]", " ", str(s or "").lower()).split()) - STOPWORDS


def same_org(wanted, candidate):
    """Guard against Discover handing back a plausible-looking wrong company."""
    a, b = name_tokens(wanted), name_tokens(candidate)
    return bool(a) and bool(b) and len(a & b) / len(a) >= 0.6


def hunter_discover(keys, org, verbose=False):
    """Company name -> domain, FREE. Discover costs no credits.

    Uses the structured 'organization' filter rather than the natural-language
    'query', because the free plan caps AI Discover searches at 10 a month
    while filter searches are uncapped. The filter schema differs across
    Hunter's docs, so try the plain form then the include-list form; a 400
    here costs nothing.
    """
    for shape in ({"organization": org}, {"organization": {"include": [org]}}):
        try:
            data = hunter(keys, "/discover", {**shape, "limit": 10}, verbose, method="post")
        except ApiError as e:
            if e.status == 400:
                continue          # wrong filter shape — try the other one
            warn("hunter discover", e)
            return ""
        for hit in (data if isinstance(data, list) else []):
            if hit.get("domain") and same_org(org, hit.get("organization")):
                return hit["domain"]
    return ""


def resolve_domain(keys, org, verbose=False):
    """Company name -> domain. Returns (domain, hunter_payload_to_reuse).

    Order matters, and it is an order of price:
      1. Discover        — free
      2. Domain Search   — 1 credit, but hands back the emails too, so when it
                           answers we keep the payload instead of paying twice
      3. Apollo          — only if Hunter drew a blank
    """
    if keys.live("hunter"):
        d = hunter_discover(keys, org, verbose)
        if d:
            return d, None
        try:
            data = hunter(keys, "/domain-search",
                          {"company": org, "limit": HUNTER_PAGE}, verbose)
            if data.get("domain"):
                return data["domain"], data
        except ApiError as e:
            warn("hunter name lookup", e)
    if keys.live("apollo"):
        try:
            res = apollo(keys, "/mixed_companies/search", {}, verbose,
                         params={"q_organization_name": org, "page": 1, "per_page": 1})
            orgs = res.get("organizations") or res.get("accounts") or []
            if orgs:
                d = clean_domain(orgs[0].get("primary_domain")
                                 or orgs[0].get("website_url") or "")
                if d:
                    return d, None
        except ApiError as e:
            warn("apollo company lookup", e)
    return "", None


# --------------------------------------------------------------------------- #
#  One row
# --------------------------------------------------------------------------- #
def clean_domain(raw):
    d = str(raw or "").strip().lower()
    for junk in ("https://", "http://", "www."):
        d = d.replace(junk, "")
    return d.split("/")[0].strip()


def lookup(keys, row, cfg, verbose=False):
    org = str(row.get("org") or "").strip()
    domain = clean_domain(row.get("domain"))
    first = str(row.get("first") or "").strip()
    last = str(row.get("last") or "").strip()
    role = str(row.get("role") or "").strip()
    mode = str(row.get("mode") or cfg["default_mode"]).strip().lower()
    if mode == "auto":
        mode = "person" if (first and last) else "domain"

    cached = None

    if not domain and org:
        domain, cached = resolve_domain(keys, org, verbose)
    if not domain:
        return blank(status="no_domain",
                     note="no provider knew this name — paste the website into column C")

    # -- a specific person
    if mode == "person" and first and last:
        if keys.live("hunter"):
            try:
                d = hunter(keys, "/email-finder",
                           {"domain": domain, "first_name": first, "last_name": last}, verbose)
                if d.get("email"):
                    conf = d.get("score") or 0
                    return {"status": "found" if conf >= cfg["min_confidence"] else "found_low_confidence",
                            "provider": "hunter", "confidence": conf, "email": d["email"],
                            "name": f"{first} {last}", "position": d.get("position") or role,
                            "domain": domain, "other": ""}
            except ApiError as e:
                warn("hunter email-finder", e)
        if keys.live("apollo"):
            try:
                p = (apollo(keys, "/people/match",
                            {"first_name": first, "last_name": last, "domain": domain,
                             "organization_name": org}, verbose) or {}).get("person") or {}
                em = p.get("email") or ""
                if em and "not_unlocked" not in em and "email_not" not in em:
                    return {"status": "found", "provider": "apollo", "confidence": "",
                            "email": em, "name": f"{first} {last}",
                            "position": p.get("title") or role, "domain": domain, "other": ""}
            except ApiError as e:
                warn("apollo people/match", e)
        mode = "domain"          # fall back to whoever we can reach there

    # -- anyone at the company
    if cached or keys.live("hunter"):
        try:
            d = cached if cached is not None else hunter(
                keys, "/domain-search", {"domain": domain, "limit": HUNTER_PAGE}, verbose)
            top, rest = best_email(d.get("emails") or [], role)
            if top and top.get("value"):
                conf = top.get("confidence") or 0
                name = f"{top.get('first_name') or ''} {top.get('last_name') or ''}".strip()
                return {"status": "found" if conf >= cfg["min_confidence"] else "found_low_confidence",
                        "provider": "hunter", "confidence": conf, "email": top["value"],
                        "name": name, "position": top.get("position") or "",
                        "domain": domain, "other": "; ".join(rest[: cfg["max_emails"]])}
        except ApiError as e:
            warn("hunter domain-search", e)

    if keys.live("apollo"):
        try:
            # Apollo splits this into two calls: api_search is free but returns
            # NO email addresses, so the best candidate then has to be enriched
            # through people/match, which is what actually costs credits.
            res = apollo(keys, "/mixed_people/api_search", {}, verbose,
                         params={"q_organization_domains_list[]": domain,
                                 "page": 1, "per_page": max(cfg["max_emails"], 5)})
            people = res.get("people") or []
            if role:
                people.sort(key=lambda p: role.lower() in str(p.get("title") or "").lower(),
                            reverse=True)
            for p in people[:1]:                     # enrich one, not the whole page
                if not p.get("id"):
                    continue
                m = (apollo(keys, "/people/match", {"id": p["id"]}, verbose) or {})
                person = m.get("person") or {}
                em = person.get("email") or ""
                if em and "not_unlocked" not in em and "email_not" not in em:
                    return {"status": "found", "provider": "apollo", "confidence": "",
                            "email": em,
                            "name": f"{person.get('first_name') or ''} "
                                    f"{person.get('last_name') or ''}".strip(),
                            "position": person.get("title") or "", "domain": domain,
                            "other": ""}
        except ApiError as e:
            warn("apollo search", e)

    return blank(domain, "not_found", f"try https://{domain}/contact")


def dry_lookup(row):
    org = str(row.get("org") or "").strip()
    domain = clean_domain(row.get("domain")) or (org.lower().replace(" ", "") + ".com.au")
    first, last = str(row.get("first") or "").strip(), str(row.get("last") or "").strip()
    if first and last:
        return {"status": "found", "provider": "dry-run", "confidence": 88,
                "email": f"{first}.{last}@{domain}".lower(), "name": f"{first} {last}",
                "position": row.get("role") or "", "domain": domain, "other": ""}
    return {"status": "found", "provider": "dry-run", "confidence": 88,
            "email": f"info@{domain}", "name": "", "position": row.get("role") or "",
            "domain": domain, "other": ""}


# --------------------------------------------------------------------------- #
#  Workbook
# --------------------------------------------------------------------------- #
def read_settings(ws):
    cfg = {"default_mode": "auto", "max_emails": 5, "min_confidence": 50,
           "delay": 0.3, "only_empty": True}
    numeric = {"max_emails_per_company": ("max_emails", int),
               "min_confidence": ("min_confidence", float),
               "rate_limit_delay_seconds": ("delay", float)}
    for r in range(1, ws.max_row + 1):
        k, v = ws.cell(r, 1).value, ws.cell(r, 2).value
        if not k or v is None:
            continue
        k = str(k).strip().lower()
        if k in numeric:
            field, cast = numeric[k]
            try:
                cfg[field] = cast(v)
            except (TypeError, ValueError):
                pass
        elif k == "default_mode":
            cfg["default_mode"] = str(v).strip().lower()
        elif k == "only_empty_rows":
            cfg["only_empty"] = str(v).strip().lower() in ("yes", "true", "1", "y")
    return cfg


def read_keys(ws):
    out = []
    for r in range(3, ws.max_row + 1):
        provider, value = ws.cell(r, 1).value, ws.cell(r, 2).value
        if not provider or not value:
            continue
        if str(ws.cell(r, 4).value).strip().lower() in ("no", "false", "0", "n"):
            continue
        if str(value).strip().upper().startswith("PASTE"):
            continue
        out.append(Key(provider, value, ws.cell(r, 3).value, ws.cell(r, 5).value))
    return out


def save(wb, path):
    try:
        wb.save(path)
        return path
    except PermissionError:
        alt = path.replace(".xlsx", "_results.xlsx")
        wb.save(alt)
        print(f"  ({path} is open in Excel — wrote {alt} instead)")
        return alt


def main():
    ap = argparse.ArgumentParser(description="Bulk work-email finder driven by an Excel workbook.")
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--check-keys", action="store_true")
    ap.add_argument("--redo", action="store_true")
    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args()

    if not args.dry_run and requests is None:
        sys.exit("Needs the 'requests' package:  pip install requests")

    wb = load_workbook(args.workbook)
    for name in (SHEET_COMPANIES, SHEET_KEYS, SHEET_SETTINGS):
        if name not in wb.sheetnames:
            sys.exit(f"Workbook has no '{name}' sheet.")
    ws = wb[SHEET_COMPANIES]
    cfg = read_settings(wb[SHEET_SETTINGS])

    keys = Keys([] if args.dry_run else read_keys(wb[SHEET_KEYS]))
    if args.check_keys:
        if not keys.all:
            sys.exit("No keys on the 'API Keys' sheet.")
        print("Checking keys (free, uses no credits):")
        check_keys(keys, args.verbose)
        return
    if not args.dry_run:
        if not keys.all:
            sys.exit("No usable keys on the 'API Keys' sheet. Add one, or use --dry-run.")
        print(f"Keys loaded: { {p: len(v) for p, v in keys.by_provider().items()} }")

    processed = found = 0
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        org = ws.cell(r, COL["org"]).value
        if not org:
            continue
        if "EXAMPLE" in f"{ws.cell(r, COL['id']).value} {org}".upper():
            continue
        if cfg["only_empty"] and not args.redo and ws.cell(r, COL["out_email"]).value:
            continue

        row = {k: ws.cell(r, COL[k]).value for k in
               ("id", "org", "domain", "first", "last", "role", "mode")}
        print(f"[row {r}] {str(org).strip()}")

        if not args.dry_run and all(k.dead for k in keys.all):
            print("  Every key has been retired. Stopping — run --check-keys to see why.")
            break

        res = dry_lookup(row) if args.dry_run else lookup(keys, row, cfg, args.verbose)

        for field, col in (("domain", "out_domain"), ("email", "out_email"),
                           ("name", "out_name"), ("position", "out_position"),
                           ("confidence", "out_confidence"), ("provider", "out_provider"),
                           ("status", "out_status"), ("other", "out_other")):
            ws.cell(r, COL[col]).value = res[field]
        ws.cell(r, COL["out_run"]).value = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d %H:%M")
        print(f"        {res['status']}  {res['email'] or res['other']}")

        processed += 1
        found += bool(res["email"])
        save(wb, args.workbook)
        if args.limit and processed >= args.limit:
            break
        if not args.dry_run:
            time.sleep(cfg["delay"])

    out = save(wb, args.workbook)
    print(f"\n{processed} rows processed, {found} emails found. Saved to {out}")
    for prov, ks in keys.by_provider().items():
        print(f"  {prov}: {sum(1 for k in ks if not k.dead)}/{len(ks)} keys live, "
              f"{sum(k.calls for k in ks)} calls")


if __name__ == "__main__":
    main()
